#!/usr/bin/env python3
"""Build an Excel workbook (Client / Competitors / Pairs sheets) from a
findings.json file already processed by compute.py.

This is a snapshot report, not an editable model: views_per_day,
like_rate, and comment_rate are written as the plain values already
computed in findings.json, not as Excel formulas.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_FINDINGS_PATH = Path(__file__).resolve().parent.parent / "findings.json"
DEFAULT_WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "workbook.xlsx"

FONT_NAME = "Arial"
BODY_FONT = Font(name=FONT_NAME, size=11)
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
LABEL_FONT = Font(name=FONT_NAME, size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)

INT_FORMAT = "#,##0"
RATE_FORMAT = "0.00"
VPD_FORMAT = "0.0"


def load_findings(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def write_table(ws, start_row, headers, rows, formats=None):
    """Writes a header row (styled) then data rows starting at start_row.
    formats is a list, same length as headers, giving a number_format
    (or None) per column. Returns the row index after the table.
    """
    formats = formats or [None] * len(headers)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=start_row + 1):
        for col, (value, fmt) in enumerate(zip(row, formats), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font = BODY_FONT
            if fmt:
                cell.number_format = fmt

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(rows) + 1


def autosize_columns(ws, widths):
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = min(max(width, 10), 60)


def video_row(video):
    return (
        video.get("title", ""),
        video.get("published_at", "")[:10] if video.get("published_at") else "",
        video.get("views"),
        video.get("views_per_day"),
        video.get("like_rate"),
        video.get("comment_rate"),
        video.get("likes"),
        video.get("comments"),
        video.get("data_source", ""),
    )


VIDEO_HEADERS = [
    "Title", "Published", "Views", "Views/day", "Like rate (%)",
    "Comment rate (%)", "Likes", "Comments", "Data source",
]
VIDEO_FORMATS = [
    None, None, INT_FORMAT, VPD_FORMAT, RATE_FORMAT, RATE_FORMAT, INT_FORMAT, INT_FORMAT, None,
]
VIDEO_WIDTHS = [50, 12, 10, 11, 14, 16, 8, 10, 13]


def build_client_sheet(wb, findings):
    ws = wb.active
    ws.title = "Client"

    client = findings.get("client") or {}
    ws.cell(row=1, column=1, value="Client").font = TITLE_FONT

    labels = [
        ("Name", client.get("name")),
        ("Channel ID", client.get("channel_id")),
        ("Subscribers", client.get("subscribers")),
        ("Capture date", client.get("capture_date")),
    ]
    row = 3
    for label, value in labels:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    row += 1
    videos = findings.get("client_videos", [])
    write_table(ws, row, VIDEO_HEADERS, [video_row(v) for v in videos], VIDEO_FORMATS)
    autosize_columns(ws, VIDEO_WIDTHS)


def build_competitors_sheet(wb, findings):
    ws = wb.create_sheet("Competitors")
    ws.cell(row=1, column=1, value="Competitors").font = TITLE_FONT

    headers = ["Channel"] + VIDEO_HEADERS
    formats = [None] + VIDEO_FORMATS
    rows = []
    for competitor in findings.get("competitors", []):
        channel_name = competitor.get("channel_name", "")
        for video in competitor.get("videos", []):
            rows.append((channel_name,) + video_row(video))

    write_table(ws, 3, headers, rows, formats)
    autosize_columns(ws, [20] + VIDEO_WIDTHS)


def build_pairs_sheet(wb, findings):
    ws = wb.create_sheet("Pairs")
    ws.cell(row=1, column=1, value="Pairs").font = TITLE_FONT

    lookup = {}
    for video in findings.get("client_videos", []):
        lookup[video.get("video_id")] = video.get("title", video.get("video_id"))
    for competitor in findings.get("competitors", []):
        for video in competitor.get("videos", []):
            lookup[video.get("video_id")] = video.get("title", video.get("video_id"))

    headers = [
        "Label", "Videos", "Held constant", "Differs",
        "Diagnosis", "Confidence", "Notes",
    ]
    rows = []
    for pair in findings.get("pairs", []):
        refs = pair.get("video_refs", [])
        videos = ", ".join(lookup.get(ref, ref) for ref in refs)
        rows.append((
            pair.get("label", ""),
            videos,
            ", ".join(pair.get("held_constant") or []),
            ", ".join(pair.get("differs") or []),
            pair.get("diagnosis", ""),
            pair.get("confidence", ""),
            pair.get("notes", ""),
        ))

    write_table(ws, 3, headers, rows)
    autosize_columns(ws, [30, 45, 25, 25, 16, 12, 50])


def main():
    findings_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FINDINGS_PATH
    workbook_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_WORKBOOK_PATH

    findings = load_findings(findings_path)

    wb = Workbook()
    build_client_sheet(wb, findings)
    build_competitors_sheet(wb, findings)
    build_pairs_sheet(wb, findings)
    wb.save(workbook_path)

    print(f"Wrote workbook to {workbook_path}.")


if __name__ == "__main__":
    main()
